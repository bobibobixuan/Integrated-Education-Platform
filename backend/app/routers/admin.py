from datetime import datetime, timedelta, timezone
from backend.app.utils import utcnow
from io import BytesIO
from pathlib import PurePath

from fastapi import Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel as _BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import case, desc, func, text

import openpyxl

from backend.app.deps import get_db, get_admin_user
from backend.app.models.user import User
from backend.app.models.unit import Unit, Level
from backend.app.models.question import Question
from backend.app.models.record import AnswerRecord, UserStats, LevelProgress
from backend.app.services.auth_service import hash_password
from backend.app.services.question_service import (
    get_effective_question_type,
    build_question_title,
    normalize_question_options,
    normalize_question_type,
)
from backend.app.services.scoring_service import calc_total_power
from backend.app.schemas.admin import AdminUserOut, AdminUserUpdate, QuestionCreate, QuestionUpdate
from backend.app.routers import admin_router


def _ensure_app_settings_table(db: Session) -> None:
    db.execute(text(
        "CREATE TABLE IF NOT EXISTS app_settings ("
        "key VARCHAR(100) PRIMARY KEY,"
        "value TEXT NOT NULL,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ")"
    ))
    db.execute(text(
        "INSERT OR IGNORE INTO app_settings(key, value) VALUES ('allow_self_register', 'true')"
    ))
    db.commit()


class _StudentImportItem(_BaseModel):
    username: str
    nickname: str
    password: str


class _ImportStudentsBody(_BaseModel):
    students: list[_StudentImportItem]


class _OptionItem(_BaseModel):
    letter: str
    text: str


class _QuestionImportItem(_BaseModel):
    title: str = ""
    type: str
    content: str
    options: list[_OptionItem] | None = None
    answer: str
    explanation: str = ""
    difficulty: int = 1


class _LevelImportItem(_BaseModel):
    name: str
    questions: list[_QuestionImportItem]


class _ImportPayload(_BaseModel):
    version: str = "1.0"
    unit: str
    questions: list[_QuestionImportItem] = Field(default_factory=list)
    levels: list[_LevelImportItem] = Field(default_factory=list)


def _validate_import_question(q: _QuestionImportItem, index_text: str) -> tuple[str, list[dict[str, str]] | None]:
    try:
        question_type = normalize_question_type(q.type)
        options = normalize_question_options(
            [{"letter": item.letter, "text": item.text} for item in (q.options or [])],
            question_type,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{index_text}: {exc}",
        ) from exc

    if not q.answer or not q.answer.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{index_text}: 缺少答案",
        )
    return question_type, options


def _flatten_import_levels(payload: _ImportPayload) -> list[tuple[str, list[_QuestionImportItem]]]:
    if payload.levels:
        return [(item.name.strip(), item.questions) for item in payload.levels if item.questions]

    grouped: list[tuple[str, list[_QuestionImportItem]]] = []
    for batch_start in range(0, len(payload.questions), 5):
        batch = payload.questions[batch_start:batch_start + 5]
        grouped.append((f"第{len(grouped) + 1}关", batch))
    return grouped


@admin_router.get("/users", response_model=list[AdminUserOut])
def list_users(db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    users = db.query(User).all()
    return [AdminUserOut.model_validate(u) for u in users]


@admin_router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    user_count = db.query(User).filter(User.role == "user").count()
    active_student_count = db.query(User).filter(User.role == "user", User.is_active == True).count()
    disabled_student_count = db.query(User).filter(User.role == "user", User.is_active == False).count()
    question_count = db.query(Question).filter(Question.is_active == True).count()
    answer_count = db.query(AnswerRecord).count()

    totals = db.query(
        func.coalesce(func.sum(UserStats.total_questions), 0),
        func.coalesce(func.sum(UserStats.total_correct), 0),
        func.coalesce(func.sum(UserStats.total_score), 0),
        func.coalesce(func.sum(UserStats.power_score), 0),
    ).one()
    total_questions, total_correct, total_score, total_power_score = [int(v or 0) for v in totals]
    avg_accuracy = round(total_correct / total_questions * 100, 1) if total_questions > 0 else 0.0

    scoring_rows = (
        db.query(UserStats.total_score)
        .filter(UserStats.total_score > 0)
        .all()
    )
    score_values = [int(row[0] or 0) for row in scoring_rows]
    avg_score = round(sum(score_values) / len(score_values), 1) if score_values else 0.0

    power_rows = (
        db.query(UserStats.power_score)
        .filter(UserStats.power_score > 0)
        .all()
    )
    power_values = [int(row[0] or 0) for row in power_rows]
    avg_power_score = round(sum(power_values) / len(power_values), 1) if power_values else 0.0

    # Last 1 hour trend, grouped by 5-minute intervals
    one_hour_ago = utcnow() - timedelta(hours=1)
    hourly_rows = (
        db.query(
            func.strftime("%H:%M", AnswerRecord.created_at).label("time_bucket"),
            func.count(AnswerRecord.id).label("count"),
        )
        .filter(AnswerRecord.created_at >= one_hour_ago)
        .group_by(func.strftime("%H:%M", AnswerRecord.created_at))
        .order_by("time_bucket")
        .all()
    )
    hourly_trend = [
        {"date": str(row.time_bucket), "count": int(row.count or 0)}
        for row in hourly_rows
    ]

    unit_rows = (
        db.query(
            Unit.name.label("unit_name"),
            func.coalesce(
                func.sum(case((AnswerRecord.is_correct == True, 1), else_=0)) * 100.0
                / func.nullif(func.count(AnswerRecord.id), 0),
                0,
            ).label("accuracy"),
            func.count(AnswerRecord.id).label("answer_count"),
            func.coalesce(func.avg(AnswerRecord.time_spent), 0).label("avg_time_spent"),
        )
        .join(Level, Level.unit_id == Unit.id)
        .join(Question, Question.level_id == Level.id)
        .join(AnswerRecord, AnswerRecord.question_id == Question.id)
        .group_by(Unit.id, Unit.name, Unit.sort_order)
        .order_by(Unit.sort_order)
        .all()
    )
    unit_accuracy = [
        {"unit_name": row.unit_name, "accuracy": round(float(row.accuracy or 0), 1)}
        for row in unit_rows
    ]

    weakest_units = [
        {
            "unit_name": row.unit_name,
            "accuracy": round(float(row.accuracy or 0), 1),
            "wrong_rate": round(100 - float(row.accuracy or 0), 1),
            "avg_time_spent": round(float(row.avg_time_spent or 0), 1),
            "answer_count": int(row.answer_count or 0),
        }
        for row in sorted(unit_rows, key=lambda item: (float(item.accuracy or 0), -int(item.answer_count or 0)))[:4]
    ]

    top_students_rows = (
        db.query(User, UserStats)
        .join(UserStats, UserStats.user_id == User.id)
        .filter(User.role == "user", User.is_active == True)
        .order_by(desc(UserStats.power_score), desc(UserStats.total_score), User.id.asc())
        .limit(5)
        .all()
    )
    top_students = [
        {
            "user_id": user.id,
            "nickname": user.nickname,
            "power_score": int(stats.power_score or 0),
            "accuracy": round((stats.total_correct or 0) / (stats.total_questions or 1) * 100, 1) if (stats.total_questions or 0) > 0 else 0.0,
            "total_questions": int(stats.total_questions or 0),
        }
        for user, stats in top_students_rows
    ]

    inactive_rows = (
        db.query(
            User.id,
            User.nickname,
            func.max(AnswerRecord.created_at).label("last_active"),
            func.coalesce(UserStats.total_questions, 0).label("total_questions"),
        )
        .outerjoin(UserStats, UserStats.user_id == User.id)
        .outerjoin(AnswerRecord, AnswerRecord.user_id == User.id)
        .filter(User.role == "user", User.is_active == True)
        .group_by(User.id, User.nickname, UserStats.total_questions)
        .order_by(func.max(AnswerRecord.created_at).asc().nullsfirst(), User.id.asc())
        .limit(5)
        .all()
    )
    inactive_students = [
        {
            "user_id": row.id,
            "nickname": row.nickname,
            "last_active": row.last_active.isoformat() if row.last_active else None,
            "total_questions": int(row.total_questions or 0),
        }
        for row in inactive_rows
    ]

    completed_level_rows = db.query(func.count(LevelProgress.id)).filter(LevelProgress.stars > 0).scalar() or 0
    total_active_levels = db.query(func.count(Level.id)).filter(Level.is_active == True).scalar() or 0
    completion_base = total_active_levels * max(active_student_count, 1)
    completion_rate = round(completed_level_rows * 100.0 / completion_base, 1) if completion_base else 0.0
    wrong_rate = round(100 - avg_accuracy, 1) if answer_count > 0 else 0.0
    total_time_spent = db.query(func.coalesce(func.sum(AnswerRecord.time_spent), 0)).scalar() or 0
    avg_weekly_activity = round((float(total_time_spent) / 60) / max(active_student_count, 1), 1) if active_student_count else 0.0
    never_practiced_count = (
        db.query(User)
        .outerjoin(UserStats, UserStats.user_id == User.id)
        .filter(User.role == "user", User.is_active == True, func.coalesce(UserStats.total_questions, 0) == 0)
        .count()
    )

    return {
        "user_count": user_count,
        "active_student_count": active_student_count,
        "disabled_student_count": disabled_student_count,
        "question_count": question_count,
        "answer_count": answer_count,
        "avg_accuracy": avg_accuracy,
        "avg_score": avg_score,
        "total_power_score": int(total_power_score),
        "avg_power_score": avg_power_score,
        "avg_weekly_activity": avg_weekly_activity,
        "completion_rate": completion_rate,
        "wrong_rate": wrong_rate,
        "never_practiced_count": never_practiced_count,
        "hourly_trend": hourly_trend,
        "unit_accuracy": unit_accuracy,
        "weakest_units": weakest_units,
        "top_students": top_students,
        "inactive_students": inactive_students,
    }


@admin_router.get("/students")
def list_students(
    sort_by: str = Query("total_score"),
    order: str = Query("desc"),
    search: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    include_disabled: bool = Query(False),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    base_q = (
        db.query(User, UserStats, func.max(AnswerRecord.created_at).label("last_active"))
        .outerjoin(UserStats, UserStats.user_id == User.id)
        .outerjoin(AnswerRecord, AnswerRecord.user_id == User.id)
        .filter(User.role == "user")
    )
    if not include_disabled:
        base_q = base_q.filter(User.is_active == True)
    if search:
        base_q = base_q.filter((User.nickname.contains(search)) | (User.username.contains(search)))

    base_q = base_q.group_by(User.id)
    total = base_q.count()

    order_col = {
        "total_score": UserStats.total_score,
        "accuracy": func.coalesce(
            UserStats.total_correct * 100.0 / func.nullif(UserStats.total_questions, 0), 0
        ),
        "total_questions": UserStats.total_questions,
    }
    col = order_col.get(sort_by, UserStats.total_score)
    if order == "asc":
        col = col.asc()
    else:
        col = col.desc()

    rows = base_q.order_by(col).offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for user, stats, last_active in rows:
        total_q = stats.total_questions if stats else 0
        total_c = stats.total_correct if stats else 0
        acc = round(total_c / total_q * 100, 1) if total_q > 0 else 0.0

        completed = (
            db.query(func.count(LevelProgress.id))
            .filter(LevelProgress.user_id == user.id, LevelProgress.stars > 0)
            .scalar()
        )
        total_stars_q = (
            db.query(func.coalesce(func.sum(LevelProgress.stars), 0))
            .filter(LevelProgress.user_id == user.id)
            .scalar()
        )

        items.append({
            "user_id": user.id,
            "username": user.username,
            "nickname": user.nickname,
            "total_score": stats.total_score if stats else 0,
            "power_score": stats.power_score if stats else 0,
            "accuracy": acc,
            "total_questions": total_q,
            "completed_levels": completed or 0,
            "total_stars": total_stars_q or 0,
            "practice_count": stats.practice_count if stats else 0,
            "last_active": last_active.isoformat() if last_active else None,
            "is_active": user.is_active,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@admin_router.get("/students/{user_id}")
def student_detail(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    stats = db.query(UserStats).filter(UserStats.user_id == user_id).first()
    levels = (
        db.query(LevelProgress)
        .filter(LevelProgress.user_id == user_id)
        .all()
    )

    total, breakdown = calc_total_power(db, user_id)

    return {
        "user_id": user.id,
        "username": user.username,
        "nickname": user.nickname,
        "is_active": user.is_active,
        "total_questions": stats.total_questions if stats else 0,
        "total_correct": stats.total_correct if stats else 0,
        "total_score": stats.total_score if stats else 0,
        "power_score": total,
        "level_progress": [
            {
                "level_id": lp.level_id,
                "stars": lp.stars,
                "unlocked": lp.unlocked,
                "best_combo": lp.best_combo,
            }
            for lp in levels
        ],
        "level_breakdown": [
            {
                "level_id": b["level_id"],
                "unit_name": b["unit_name"],
                "level_name": b["level_name"],
                "clear": b["clear"],
                "perfect": b["perfect"],
                "speed": b["speed"],
                "combo": b["combo"],
                "total": b["total"],
            }
            for b in breakdown
        ],
    }


class _CreateStudentBody(_BaseModel):
    username: str
    nickname: str
    password: str


@admin_router.post("/students")
def create_student(
    body: _CreateStudentBody,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    existing = (
        db.query(User)
        .filter(User.username == body.username, User.is_active == True)
        .first()
    )
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已存在")

    if len(body.password) < 6:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="密码至少6位")

    user = User(
        username=body.username,
        original_username=body.username,
        password_hash=hash_password(body.password),
        nickname=body.nickname,
        role="user",
        force_password_change=False,
    )
    db.add(user)
    db.flush()

    stats = UserStats(user_id=user.id)
    db.add(stats)
    db.commit()
    db.refresh(user)

    return {"id": user.id, "username": user.username, "nickname": user.nickname, "message": "学生已创建"}


@admin_router.post("/students/import")
def import_students(
    body: _ImportStudentsBody,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    if not body.students:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="没有可导入的学生")
    if len(body.students) > 300:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="单次最多导入300名学生")

    usernames = [item.username.strip() for item in body.students]
    if len(set(usernames)) != len(usernames):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入列表里存在重复用户名")

    existing = {
        row[0]
        for row in db.query(User.username).filter(User.username.in_(usernames), User.is_active == True).all()
    }
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"以下用户名已存在：{', '.join(sorted(existing))}",
        )

    created = 0
    for item in body.students:
        username = item.username.strip()
        nickname = item.nickname.strip()
        password = item.password.strip()
        if not username or not nickname:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="用户名和昵称不能为空")
        if len(password) < 6:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"学生 {nickname} 的密码至少6位")

        user = User(
            username=username,
            original_username=username,
            password_hash=hash_password(password),
            nickname=nickname,
            role="user",
            force_password_change=False,
        )
        db.add(user)
        db.flush()
        db.add(UserStats(user_id=user.id))
        created += 1

    db.commit()
    return {"message": f"已成功导入 {created} 名学生", "created": created}


@admin_router.put("/students/{user_id}")
def update_student(
    user_id: int,
    body: AdminUserUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if target.role != "user":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不能修改管理员账号")

    if body.nickname is not None:
        target.nickname = body.nickname
    if body.role is not None:
        target.role = body.role
    if body.is_active is not None:
        target.is_active = body.is_active
    if body.force_password_change is not None:
        target.force_password_change = body.force_password_change
    if body.new_password is not None:
        target.password_hash = hash_password(body.new_password)
        db.execute(
            text("UPDATE users SET token_version = token_version + 1 WHERE id = :id"),
            {"id": user_id},
        )
        db.refresh(target)

    db.commit()
    return {"id": target.id, "message": "学生信息已更新"}


@admin_router.delete("/students/{user_id}")
def delete_student(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if target.role != "user":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不能禁用管理员账号")

    now = datetime.now(timezone.utc)
    ts = now.strftime("%Y%m%d%H%M%S")
    target.original_username = target.username
    target.username = f"disabled_{user_id}_{ts}"
    target.is_active = False
    target.disabled_at = now

    db.execute(
        text("UPDATE users SET token_version = token_version + 1 WHERE id = :id"),
        {"id": user_id},
    )
    db.commit()
    return {"id": target.id, "message": "学生已禁用"}


@admin_router.get("/analytics/levels")
def level_analytics(
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    rows = (
        db.query(
            Unit.name.label("unit_name"),
            Level.name.label("level_name"),
            func.count(AnswerRecord.id).label("total_attempts"),
            func.coalesce(
                func.sum(case((AnswerRecord.is_correct == True, 1), else_=0)) * 100.0
                / func.nullif(func.count(AnswerRecord.id), 0),
                0,
            ).label("correct_rate"),
            func.coalesce(func.avg(AnswerRecord.time_spent), 0).label("avg_time_spent"),
            func.count(func.distinct(AnswerRecord.user_id)).label("student_count"),
        )
        .join(Question, Question.id == AnswerRecord.question_id)
        .join(Level, Level.id == Question.level_id)
        .join(Unit, Unit.id == Level.unit_id)
        .filter(Level.is_active == True)
        .group_by(Unit.id, Unit.name, Unit.sort_order, Level.id, Level.name, Level.sort_order)
        .order_by(Unit.sort_order, Level.sort_order)
        .all()
    )

    return [
        {
            "unit_name": row.unit_name,
            "level_name": row.level_name,
            "total_attempts": int(row.total_attempts or 0),
            "correct_rate": round(float(row.correct_rate or 0), 1),
            "avg_time_spent": round(float(row.avg_time_spent or 0), 1),
            "student_count": int(row.student_count or 0),
        }
        for row in rows
    ]


@admin_router.get("/analytics/wrong-questions")
def wrong_question_stats(
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    rows = (
        db.query(
            Question.id.label("question_id"),
            Question.content.label("question_content"),
            Unit.name.label("unit_name"),
            Level.name.label("level_name"),
            func.sum(case((AnswerRecord.is_correct == False, 1), else_=0)).label("wrong_count"),
            func.count(AnswerRecord.id).label("total_attempts"),
        )
        .join(AnswerRecord, AnswerRecord.question_id == Question.id)
        .join(Level, Level.id == Question.level_id)
        .join(Unit, Unit.id == Level.unit_id)
        .group_by(Question.id, Question.content, Unit.name, Level.name)
        .having(func.sum(case((AnswerRecord.is_correct == False, 1), else_=0)) > 0)
        .order_by(desc("wrong_count"), desc("total_attempts"), Question.id.asc())
        .limit(limit)
        .all()
    )

    result = []
    for row in rows:
        total_attempts = int(row.total_attempts or 0)
        wrong_count = int(row.wrong_count or 0)
        result.append({
            "question_id": int(row.question_id),
            "question_content": row.question_content,
            "unit_name": row.unit_name,
            "level_name": row.level_name,
            "wrong_count": wrong_count,
            "wrong_rate": round(wrong_count * 100.0 / total_attempts, 1) if total_attempts else 0.0,
            "total_attempts": total_attempts,
        })
    return result


@admin_router.get("/questions")
def list_questions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    level_id: int = Query(None),
    unit_id: int = Query(None),
    question_type: str = Query(""),
    include_inactive: bool = Query(False),
    search: str = Query(""),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    base_q = (
        db.query(Question, Level.name, Unit.name, Unit.id)
        .join(Level, Level.id == Question.level_id)
        .join(Unit, Unit.id == Level.unit_id)
    )
    if not include_inactive:
        base_q = base_q.filter(Question.is_active == True)
    if unit_id is not None:
        base_q = base_q.filter(Unit.id == unit_id)
    if level_id is not None:
        base_q = base_q.filter(Question.level_id == level_id)
    if question_type:
        base_q = base_q.filter(Question.type == question_type)
    if search:
        base_q = base_q.filter((Question.content.contains(search)) | (Question.title.contains(search)))

    total = base_q.count()
    rows = (
        base_q
        .order_by(Unit.sort_order, Level.sort_order, Question.sort_order)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [
        {
            "id": q.id,
            "title": q.title or "",
            "content": q.content,
            "type": get_effective_question_type(q),
            "answer": q.answer,
            "options": q.options,
            "level_id": q.level_id,
            "level_name": level_name,
            "unit_id": unit_id_val,
            "unit_name": unit_name,
            "sort_order": q.sort_order,
            "is_active": q.is_active,
            "knowledge_meaning": q.knowledge_meaning or "",
            "knowledge_rule": q.knowledge_rule or "",
            "knowledge_error": q.knowledge_error or "",
            "knowledge_example": q.knowledge_example or "",
        }
        for q, level_name, unit_name, unit_id_val in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@admin_router.post("/questions")
def create_question(
    body: QuestionCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    level = db.query(Level).filter(Level.id == body.level_id).first()
    if not level:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关卡不存在")
    from backend.app.services.question_service import normalize_question_type, normalize_question_options
    try:
        qtype = normalize_question_type(body.type)
        opts = normalize_question_options(body.options, qtype)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    q = Question(
        level_id=body.level_id,
        title=body.title or body.content[:60],
        type=qtype,
        content=body.content.strip(),
        options=opts,
        answer=body.answer.strip(),
        knowledge_meaning=body.knowledge_meaning or "",
        knowledge_rule=body.knowledge_rule or "",
        knowledge_error=body.knowledge_error or "",
        knowledge_example=body.knowledge_example or "",
        sort_order=body.sort_order,
        is_active=True,
    )
    db.add(q)
    db.flush()
    level.questions_count = (level.questions_count or 0) + 1
    db.commit()
    db.refresh(q)
    return {"id": q.id, "message": "题目已创建"}


@admin_router.put("/questions/{question_id}")
def update_question(
    question_id: int,
    body: QuestionUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    q = db.query(Question).filter(Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="题目不存在")

    if body.content is not None:
        q.content = body.content.strip()
    if body.title is not None:
        q.title = body.title
    if body.type is not None:
        from backend.app.services.question_service import normalize_question_type, normalize_question_options
        try:
            qtype = normalize_question_type(body.type)
            q.type = qtype
            if body.options is not None:
                q.options = normalize_question_options(body.options, qtype)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    elif body.options is not None:
        q.options = body.options
    if body.answer is not None:
        q.answer = body.answer.strip()
    if body.level_id is not None:
        new_level = db.query(Level).filter(Level.id == body.level_id).first()
        if not new_level:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关卡不存在")
        q.level_id = body.level_id
    if body.is_active is not None:
        q.is_active = body.is_active
    if body.sort_order is not None:
        q.sort_order = body.sort_order
    if body.knowledge_meaning is not None:
        q.knowledge_meaning = body.knowledge_meaning
    if body.knowledge_rule is not None:
        q.knowledge_rule = body.knowledge_rule
    if body.knowledge_error is not None:
        q.knowledge_error = body.knowledge_error
    if body.knowledge_example is not None:
        q.knowledge_example = body.knowledge_example

    db.commit()
    return {"id": q.id, "message": "题目已更新"}


@admin_router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    q = db.query(Question).filter(Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="题目不存在")
    q.is_active = False
    level = db.query(Level).filter(Level.id == q.level_id).first()
    if level:
        level.questions_count = max(0, (level.questions_count or 1) - 1)
    db.commit()
    return {"id": q.id, "message": "题目已禁用"}


@admin_router.get("/settings/registration")
def get_registration_setting(
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    _ensure_app_settings_table(db)
    row = db.execute(
        text("SELECT value FROM app_settings WHERE key = 'allow_self_register'")
    ).fetchone()
    return {"allow_self_register": str(row[0]).lower() != "false" if row else True}


@admin_router.put("/settings/registration")
def update_registration_setting(
    body: dict,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    enabled = body.get("allow_self_register")
    if not isinstance(enabled, bool):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="allow_self_register 必须是布尔值")

    _ensure_app_settings_table(db)
    db.execute(
        text(
            "INSERT INTO app_settings(key, value, updated_at) VALUES ('allow_self_register', :value, CURRENT_TIMESTAMP) "
            "ON CONFLICT(key) DO UPDATE SET value = :value, updated_at = CURRENT_TIMESTAMP"
        ),
        {"value": "true" if enabled else "false"},
    )
    db.commit()
    return {"allow_self_register": enabled}


@admin_router.post("/import")
def import_questions(
    payload: _ImportPayload,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    groups = _flatten_import_levels(payload)
    if not groups:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="需要提供 questions 或 levels 字段")

    total_questions = sum(len(items) for _, items in groups)
    if total_questions > 500:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="单次最多导入500题")

    for level_index, (level_name, questions) in enumerate(groups, start=1):
        if not level_name.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {level_index} 个分组缺少关卡名称")
        for question_index, q in enumerate(questions, start=1):
            _validate_import_question(q, f"{level_name} 第{question_index}题")

    unit = db.query(Unit).filter(Unit.name == payload.unit).first()
    if not unit:
        max_order = db.query(Unit).count()
        unit = Unit(
            name=payload.unit,
            icon="📚",
            subtitle=payload.unit,
            description=payload.unit,
            learning_goal="",
            coach_line="",
            starter_tip="",
            color="#0f766e",
            sort_order=max_order,
        )
        db.add(unit)
        db.flush()

    existing_max_sort = (
        db.query(Level)
        .filter(Level.unit_id == unit.id)
        .order_by(Level.sort_order.desc())
        .first()
    )
    base_sort = (existing_max_sort.sort_order + 1) if existing_max_sort else 0

    imported = 0
    created_levels = 0
    for group_index, (group_name, questions) in enumerate(groups):
        level = Level(
            unit_id=unit.id,
            name=group_name,
            icon="📝",
            bg="🏰",
            questions_count=len(questions),
            sort_order=base_sort + group_index,
        )
        db.add(level)
        db.flush()
        created_levels += 1

        for idx, q in enumerate(questions):
            question_type, options_list = _validate_import_question(q, f"{group_name} 第{idx + 1}题")
            question = Question(
                level_id=level.id,
                title=build_question_title(q.title, q.content),
                type=question_type,
                content=q.content.strip(),
                options=options_list,
                answer=q.answer.strip(),
                knowledge_meaning=q.explanation or "",
                knowledge_rule="",
                knowledge_error="",
                knowledge_example="",
                sort_order=idx,
                is_active=True,
            )
            db.add(question)
            imported += 1

    db.commit()
    return {"message": f"导入成功：单元「{payload.unit}」新增 {imported} 题，{created_levels} 关"}


# ── Excel 导入辅助 ──

_EXCEL_TYPE_MAP: dict[str, str] = {
    "单选题": "选择题",
    "多选题": "多选题",
    "判断题": "判断题",
    "填空题": "填空题",
}


def _read_excel_sheet(file_bytes: bytes):
    """加载 Excel 并返回活跃工作表。"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 工作簿没有任何工作表")
    return ws


def _detect_excel_format(ws) -> tuple[int, dict[str, int]]:
    """检测 Excel 格式，返回 (header_row_index, {field_name: col_index})。

    支持两种格式：
    1. 模板格式 — 第1行说明，第2行表头
    2. 题目列表格式 — 第1行表头（含"目录"列）
    """
    # 读取前两行来判断
    row1_vals = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, min(ws.max_column + 1, 20))]
    row2_vals = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, min(ws.max_column + 1, 20))]

    def _score_header_row(vals: list[str]) -> int:
        """给一行打分，匹配的关键词越多分数越高"""
        score = 0
        kw_map = {
            "题型": 3, "难度": 2, "知识点": 2, "分值": 1,
            "题目内容": 3, "答案": 3, "选项": 2, "目录": 3,
        }
        for v in vals:
            for kw, s in kw_map.items():
                if kw in v:
                    score += s
        return score

    score1 = _score_header_row(row1_vals)
    score2 = _score_header_row(row2_vals)

    if score1 >= score2 and score1 >= 5:
        header_row = 1
        header_vals = row1_vals
    elif score2 >= 5:
        header_row = 2
        header_vals = row2_vals
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="无法识别 Excel 表头，请确保包含「题型」「题目内容」「答案」等列",
        )

    # 映射列名到列索引
    col_map: dict[str, int] = {}
    option_start = 99
    for idx, val in enumerate(header_vals):
        if not val:
            continue
        if "题型" in val:
            col_map["type"] = idx
        elif "难度" in val:
            col_map["difficulty"] = idx
        elif "知识点" in val:
            col_map["knowledge"] = idx
        elif "分值" in val:
            col_map["score"] = idx
        elif "题目内容" in val:
            col_map["content"] = idx
        elif "答案" in val or val.strip() == "答":
            col_map["answer"] = idx
        elif "标题" in val:
            col_map["title"] = idx
        elif "目录" in val:
            col_map["category"] = idx
        elif "选项" in val or val.startswith("选项"):
            option_start = min(option_start, idx)
        elif "内容" in val:
            col_map["description"] = idx

    # 如果没有明确的「题目内容」列但有「内容」列，用「内容」作为题目内容
    if "content" not in col_map and "description" in col_map:
        col_map["content"] = col_map.pop("description")

    # 选项列：答案列之后的所有非空表头列
    if option_start == 99 and "answer" in col_map:
        option_start = col_map["answer"] + 1
    col_map["options_start"] = option_start

    return header_row, col_map


def _parse_excel_to_questions(file_bytes: bytes) -> tuple[list[_QuestionImportItem], str | None, str | None]:
    """解析 Excel 为题目列表。返回 (questions, auto_unit_name, knowledge_point_column)。

    knowledge_point_column 返回知识点列的原始值，用于后续分组。
    """
    ws = _read_excel_sheet(file_bytes)
    header_row, cm = _detect_excel_format(ws)

    data_start = header_row + 1
    max_row = ws.max_row
    max_col = ws.max_column or 20

    questions: list[_QuestionImportItem] = []
    knowledge_points: list[str] = []  # 与 questions 一一对应
    auto_unit: str | None = None

    # 尝试从目录列提取单元名（后备）
    if "category" in cm and data_start <= max_row:
        first_cat = ws.cell(row=data_start, column=cm["category"] + 1).value
        if first_cat and str(first_cat).strip():
            raw = str(first_cat).strip()
            parts = [p.strip() for p in raw.replace("\\", "/").split("/") if p.strip()]
            auto_unit = parts[-1] if parts else raw

    for row_idx in range(data_start, max_row + 1):
        type_val = ws.cell(row=row_idx, column=cm["type"] + 1).value if "type" in cm else None
        if type_val is None:
            continue
        qtype_raw = str(type_val).strip()
        if not qtype_raw:
            continue

        qtype_mapped = _EXCEL_TYPE_MAP.get(qtype_raw)
        if not qtype_mapped:
            continue

        def _get(field: str, default=""):
            if field not in cm:
                return default
            v = ws.cell(row=row_idx, column=cm[field] + 1).value
            return str(v).strip() if v is not None else default

        content = _get("content")
        if not content:
            content = _get("description")

        if not content:
            continue

        answer = _get("answer")
        if not answer:
            continue

        difficulty = 1
        if "difficulty" in cm:
            try:
                difficulty = int(ws.cell(row=row_idx, column=cm["difficulty"] + 1).value or 1)
            except (ValueError, TypeError):
                difficulty = 1

        knowledge = _get("knowledge")
        title = _get("title")

        # 收集选项
        option_letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        options: list[_OptionItem] = []
        opt_start = cm.get("options_start", 99)
        for j in range(opt_start, max_col):
            opt_val = ws.cell(row=row_idx, column=j + 1).value
            if opt_val is not None and str(opt_val).strip():
                letter_idx = j - opt_start
                letter = option_letters[letter_idx] if letter_idx < len(option_letters) else f"opt{letter_idx}"
                options.append(_OptionItem(letter=letter, text=str(opt_val).strip()))

        explanation_parts = []
        if knowledge:
            explanation_parts.append(f"知识点：{knowledge}")
        explanation = "；".join(explanation_parts) if explanation_parts else ""

        questions.append(_QuestionImportItem(
            title=title,
            type=qtype_mapped,
            content=content,
            options=options if options else None,
            answer=answer,
            explanation=explanation,
            difficulty=difficulty,
        ))
        knowledge_points.append(knowledge)

    if not questions:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 表格中未解析到任何有效题目")

    return questions, auto_unit, knowledge_points


def _extract_unit_name(kp: str) -> str:
    """从知识点路径中提取单元名称。取最后一层有意义的名字。"""
    if not kp:
        return "未分类"
    parts = [p.strip() for p in kp.replace("\\", "/").replace("-", "/").split("/") if p.strip()]
    # 取最后一个有意义的部分（去掉太短的如单个字母）
    for part in reversed(parts):
        clean = part.strip().rstrip("0123456789.-_ ")
        if len(clean) >= 2:
            return clean
    return parts[-1] if parts else "未分类"


def _group_questions_by_knowledge(
    questions: list[_QuestionImportItem],
    knowledge_points: list[str],
) -> dict[str, list[_QuestionImportItem]]:
    """按知识点分组，返回 {unit_name: questions}。"""
    groups: dict[str, list[_QuestionImportItem]] = {}
    for q, kp in zip(questions, knowledge_points):
        unit_name = _extract_unit_name(kp)
        groups.setdefault(unit_name, []).append(q)
    return groups


@admin_router.post("/import/excel")
async def import_questions_from_excel(
    file: UploadFile = File(...),
    unit: str = Query("", description="目标单元名称，留空则从Excel「目录」列自动提取"),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    ext = PurePath(file.filename or "unknown.xlsx").suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 .xlsx / .xlsm 格式的 Excel 文件")

    try:
        file_bytes = await file.read()
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="读取上传文件失败")

    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件大小不能超过 10 MB")

    if len(file_bytes) < 5:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空或已损坏")

    questions, auto_unit, knowledge_points = _parse_excel_to_questions(file_bytes)

    if len(questions) > 500:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="单次最多导入500题")

    # 按知识点分组：每个知识点 → 一个单元 → 一个关卡（全部题目）
    from collections import OrderedDict
    unit_groups: OrderedDict[str, list[_QuestionImportItem]] = OrderedDict()
    for q, kp in zip(questions, knowledge_points):
        unit_name = _extract_unit_name(kp)
        unit_groups.setdefault(unit_name, []).append(q)

    if len(unit_groups) > 20:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"知识点分组过多（{len(unit_groups)}个），最多支持20个单元")

    total_imported = 0
    total_units = 0
    messages: list[str] = []

    for unit_name, qs in unit_groups.items():
        # 验证题目
        for qi, q in enumerate(qs, start=1):
            _validate_import_question(q, f"「{unit_name}」第{qi}题")

        # 查找或创建单元
        unit_obj = db.query(Unit).filter(Unit.name == unit_name).first()
        if not unit_obj:
            max_order = db.query(Unit).count()
            unit_obj = Unit(
                name=unit_name,
                icon="📚",
                subtitle=unit_name,
                description=unit_name,
                learning_goal="",
                coach_line="",
                starter_tip="",
                color="#0f766e",
                sort_order=max_order,
            )
            db.add(unit_obj)
            db.flush()

        # 所有题目放入一个关卡
        level = Level(
            unit_id=unit_obj.id,
            name="全部题目",
            icon="📝",
            bg="🏰",
            questions_count=len(qs),
            sort_order=0,
        )
        db.add(level)
        db.flush()

        for idx, q in enumerate(qs):
            question_type, options_list = _validate_import_question(q, f"「{unit_name}」第{idx + 1}题")
            question = Question(
                level_id=level.id,
                title=build_question_title(q.title, q.content),
                type=question_type,
                content=q.content.strip(),
                options=options_list,
                answer=q.answer.strip(),
                knowledge_meaning=q.explanation or "",
                knowledge_rule="",
                knowledge_error="",
                knowledge_example="",
                sort_order=idx,
                is_active=True,
            )
            db.add(question)

        total_imported += len(qs)
        total_units += 1
        messages.append(f"「{unit_name}」{len(qs)}题")

    db.commit()
    return {"message": f"Excel 导入成功：{total_units} 个单元，共 {total_imported} 题（{' / '.join(messages)}）"}
